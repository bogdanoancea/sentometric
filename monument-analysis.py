#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Scrape article text from Excel list of URLs,
infer CAMEO codes (subset or fast hierarchical zero-shot, multilingual),
join Goldstein, and compute sentometrics.

- "HF" sentiment branch uses ZERO-SHOT sentiment with
  MoritzLaurer/mDeBERTa-v3-base-mnli-xnli (news-friendly)
- VADER augmentation:
  * RO/HU lexicon for vandalism/heritage domain
  * asymmetric rescaling (shrink positives, slightly amplify negatives)
  * vandalism context penalty
- CAMEO-driven domain prior for destructive/hostile families
- Asymmetric neutral window: default --sent-asym-neg 0.03 --sent-asym-pos 0.10

Install:
  pip install -U pandas openpyxl requests beautifulsoup4 trafilatura lxml \
                 transformers torch nltk langdetect sentencepiece sacremoses
"""

from __future__ import annotations
import argparse
import sys
import re
import hashlib
from typing import Optional, List, Tuple, Dict, Iterable

import pandas as pd
import requests
from bs4 import BeautifulSoup
import trafilatura
from trafilatura.settings import use_config

import torch
from transformers import pipeline, AutoTokenizer, AutoModelForSeq2SeqLM

# NLTK for VADER + sentence split
import nltk
from nltk.sentiment import SentimentIntensityAnalyzer
from nltk.tokenize import sent_tokenize

import numpy as np

# Optional language detection (for VADER translation auto-selection)
try:
    from langdetect import detect, DetectorFactory
    DetectorFactory.seed = 42
    HAVE_LANGDETECT = True
except Exception:
    HAVE_LANGDETECT = False

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
      "AppleWebKit/537.36 (KHTML, like Gecko) "
      "Chrome/124.0 Safari/537.36")
MIN_WORDS = 8

# ---------------------------
# Excel URL utilities
# ---------------------------
def looks_like_url(x: str) -> bool:
    if not isinstance(x, str):
        return False
    x = x.strip()
    return x.startswith(("http://", "https://", "www."))

def normalize_url(u: str) -> str:
    u = (u or "").strip()
    return "http://" + u if u.startswith("www.") else u

def read_urls_with_hyperlinks(
    xlsx_path: str, sheet: Optional[str], url_col: int, one_based: bool
) -> Tuple[List[str], List[int], pd.DataFrame]:
    """
    Read URLs from a specific column (no autodetect).
    Uses hyperlink target when present; otherwise visible cell text.
    Works around openpyxl ReadOnlyCell (no .hyperlink) by trying normal mode first.
    Returns (urls, kept_row_indices, df_values_for_alignment).
    """
    from openpyxl import load_workbook

    def _collect(ws, use_hyperlinks: bool):
        rows_vals, rows_hrefs = [], []
        for row in ws.iter_rows(values_only=False):
            vals, hrefs = [], []
            for cell in row:
                vals.append("" if cell.value is None else str(cell.value))
                if use_hyperlinks and getattr(cell, "hyperlink", None) and cell.hyperlink.target:
                    hrefs.append(str(cell.hyperlink.target))
                else:
                    hrefs.append("")
            rows_vals.append(vals)
            rows_hrefs.append(hrefs)
        return rows_vals, rows_hrefs

    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=False, keep_links=True)
        ws = wb[sheet] if sheet else wb.active
        rows_vals, rows_hrefs = _collect(ws, use_hyperlinks=True)
    except Exception:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows_vals, rows_hrefs = _collect(ws, use_hyperlinks=False)

    if not rows_vals:
        return [], [], pd.DataFrame()

    df_vals = pd.DataFrame(rows_vals)
    df_hrefs = pd.DataFrame(rows_hrefs)

    col_idx = url_col - 1 if one_based else url_col
    if col_idx < 0 or col_idx >= df_vals.shape[1]:
        raise ValueError("url_col out of range.")

    urls, idxs = [], []
    for i in range(df_vals.shape[0]):
        href = df_hrefs.iat[i, col_idx] if col_idx < df_hrefs.shape[1] else ""
        text = df_vals.iat[i, col_idx]
        candidate = href if looks_like_url(href) else text
        if looks_like_url(candidate):
            urls.append(normalize_url(candidate))
            idxs.append(i)

    print(f"[info] extracted {len(urls)} URL(s) from column {url_col} ({'1-based' if one_based else '0-based'})")
    return urls, idxs, df_vals

# ---------------------------
# Web fetch + text extraction
# ---------------------------
def build_trafilatura_config():
    cfg = use_config()
    cfg.set("DEFAULT", "EXTRACTION_TIMEOUT", "120")
    cfg.set("DEFAULT", "MIN_EXTRACTED_SIZE", "10")
    cfg.set("DEFAULT", "FAVOR_RECALL", "true")
    cfg.set("DEFAULT", "STRICT", "false")
    cfg.set("DEFAULT", "USER_AGENT", UA)
    return cfg

def fetch_html(url: str, timeout: int = 40) -> Tuple[Optional[int], str]:
    try:
        r = requests.get(url, headers={"User-Agent": UA}, timeout=timeout, allow_redirects=True)
        status = r.status_code
        if status >= 400 or not r.content:
            return status, ""
        enc = r.encoding or r.apparent_encoding or "utf-8"
        try:
            html = r.content.decode(enc, errors="replace")
        except Exception:
            html = r.text
        return status, html.replace("\x00", " ")
    except Exception:
        return None, ""

def extract_text_and_title(html: str) -> Tuple[str, str]:
    """
    Prefer trafilatura; fallback to concatenated <p> nodes if short/empty.
    """
    text, title = "", ""
    if html:
        try:
            text = trafilatura.extract(
                html, output_format="txt",
                include_comments=False, include_tables=False,
                favor_recall=True, no_fallback=False, with_metadata=False
            ) or ""
        except Exception:
            text = ""
        if len(text.split()) < MIN_WORDS:
            soup = BeautifulSoup(html, "lxml")
            node = soup.find("article") or soup
            text = " ".join(p.get_text(" ", strip=True) for p in node.find_all("p"))
        try:
            soup2 = BeautifulSoup(html, "lxml")
            if soup2.title and soup2.title.string:
                title = soup2.title.string.strip()
        except Exception:
            title = ""
    return text.strip(), title

def scrape_urls(urls: List[str], timeout: int) -> pd.DataFrame:
    """
    Fetch + extract for all URLs. Returns DataFrame with text and basic meta.
    """
    records = []
    for i, url in enumerate(urls, start=1):
        status, html = fetch_html(url, timeout=timeout)
        if not html:
            records.append(dict(id=i, url=url, http_status=status, title=None,
                                text_chars=0, text="", error="fetch-failed"))
            print(f"[{i}/{len(urls)}] fetch-failed {status} {url}")
            continue
        text, title = extract_text_and_title(html)
        if not text:
            records.append(dict(id=i, url=url, http_status=status, title=title or None,
                                text_chars=0, text="", error="no-text"))
            print(f"[{i}/{len(urls)}] no-text {url}")
            continue
        records.append(dict(id=i, url=url, http_status=status, title=title or None,
                            text_chars=len(text), text=text, error=None))
        print(f"[{i}/{len(urls)}] ok {url} ({len(text)} chars)")
    return pd.DataFrame.from_records(records, columns=["id","url","http_status","title","text_chars","text","error"])

# ---------------------------
# CAMEO tables + zero-shot
# ---------------------------
def read_mapping_any(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, sep=None, engine="python", dtype=str).fillna("")
    df.columns = [c.strip().upper() for c in df.columns]
    return df

def load_cameo_tables(eventcodes_path: str, goldstein_path: str) -> Tuple[pd.DataFrame, Dict[str,str], Dict[str,float]]:
    ev_df = read_mapping_any(eventcodes_path)
    gs_df = read_mapping_any(goldstein_path)

    if "EVENTDESCRIPTION" not in ev_df.columns and "DESCRIPTION" in ev_df.columns:
        ev_df = ev_df.rename(columns={"DESCRIPTION": "EVENTDESCRIPTION"})
    if "GOLDSTEINSCALE" not in gs_df.columns and "GOLDSTEIN" in gs_df.columns:
        gs_df = gs_df.rename(columns={"GOLDSTEIN": "GOLDSTEINSCALE"})

    required_ev = {"CAMEOEVENTCODE", "EVENTDESCRIPTION"}
    if not required_ev.issubset(set(ev_df.columns)):
        raise ValueError(f"EVENTCODES missing columns: {required_ev - set(ev_df.columns)}")
    required_gs = {"CAMEOEVENTCODE", "GOLDSTEINSCALE"}
    if not required_gs.issubset(set(gs_df.columns)):
        raise ValueError(f"GOLDSTEIN missing columns: {required_gs - set(gs_df.columns)}")

    ev_df["CAMEOEVENTCODE"] = ev_df["CAMEOEVENTCODE"].astype(str).str.strip()
    gs_df["CAMEOEVENTCODE"] = gs_df["CAMEOEVENTCODE"].astype(str).str.strip()

    ev_desc = dict(zip(ev_df["CAMEOEVENTCODE"], ev_df["EVENTDESCRIPTION"]))
    gs_map: Dict[str, float] = {}
    for k, v in zip(gs_df["CAMEOEVENTCODE"], gs_df["GOLDSTEINSCALE"]):
        try:
            gs_map[k] = float(v)
        except Exception:
            pass

    return ev_df, ev_desc, gs_map

def lookup_with_prefix_fallback(code: str, ev_desc: Dict[str, str], gs_map: Dict[str, float]) -> Tuple[str, float]:
    code = (code or "").strip()
    if not code:
        return "(unknown)", 0.0
    desc = ev_desc.get(code, "")
    gs = gs_map.get(code)
    k = code
    while (not desc or gs is None) and len(k) > 1:
        k = k[:-1]
        desc = desc or ev_desc.get(k, "")
        gs = gs if gs is not None else gs_map.get(k)
    return (desc if desc else "(unknown)"), (gs if gs is not None else 0.0)

def descendant_codes(ev_df: pd.DataFrame, parent_code: str, max_depth_extra: int, max_candidates: int) -> List[Tuple[str, str]]:
    plen = len(parent_code)
    max_len = plen + max_depth_extra
    sub = ev_df[
        ev_df["CAMEOEVENTCODE"].str.startswith(parent_code)
        & (ev_df["CAMEOEVENTCODE"].str.len() > plen)
        & (ev_df["CAMEOEVENTCODE"].str.len() <= max_len)
    ].sort_values("CAMEOEVENTCODE")
    pairs = list(zip(sub["CAMEOEVENTCODE"], sub["EVENTDESCRIPTION"]))
    if not pairs:
        pairs = [(parent_code, "")]
    return pairs[:max_candidates]

def build_zs(model_name: str, device: int):
    return pipeline("zero-shot-classification", model=model_name, device=device)

def _zs_best_label(text: str, labels: List[str], zs_pipe, hypothesis_template: Optional[str]):
    if not labels:
        return None, 0.0
    res = zs_pipe(
        text,
        candidate_labels=labels,
        hypothesis_template=hypothesis_template or "Această propoziție este despre {}.",
        multi_label=False,
    )
    return res["labels"][0], float(res["scores"][0])

def coarse_pick_root(text: str, coarse_codes: List[str], ev_df: pd.DataFrame, zs_pipe, hypothesis_template: Optional[str]):
    pairs = []
    for c in coarse_codes:
        sub = ev_df.loc[ev_df["CAMEOEVENTCODE"] == c, "EVENTDESCRIPTION"]
        desc = sub.values[0] if len(sub) else ""
        pairs.append((c, desc))
    labels = [f"[{c}] {d}" if d else f"[{c}]" for c, d in pairs]
    lab, scr = _zs_best_label(text, labels, zs_pipe, hypothesis_template)
    if not lab:
        return None, 0.0
    m = re.match(r"\[(.*?)\]", lab)
    return (m.group(1) if m else None), scr

def _normalize_kw_list(s: str) -> List[str]:
    if not s or not s.strip():
        return []
    return [x.strip().lower() for x in s.split(",") if x.strip()]

def subset_from_keywords_and_seeds(
    ev_df: pd.DataFrame,
    seed_codes: Iterable[str],
    keywords: Iterable[str],
    depth: int,
    max_size: int
) -> List[Tuple[str,str]]:
    codes_set = set()
    # keyword matches
    if keywords:
        descs = ev_df[["CAMEOEVENTCODE", "EVENTDESCRIPTION"]].dropna()
        for _, r in descs.iterrows():
            code = r["CAMEOEVENTCODE"]
            desc = str(r["EVENTDESCRIPTION"]).lower()
            if any(kw in desc for kw in keywords):
                codes_set.add(code)

    # seed descendants
    for sc in (seed_codes or []):
        sc = sc.strip()
        if not sc:
            continue
        matches = descendant_codes(ev_df, sc, max_depth_extra=depth, max_candidates=10_000)
        for c, _ in matches:
            codes_set.add(c)
        codes_set.add(sc)

    sub = ev_df[ev_df["CAMEOEVENTCODE"].isin(sorted(codes_set))].sort_values("CAMEOEVENTCODE")
    pairs = list(zip(sub["CAMEOEVENTCODE"], sub["EVENTDESCRIPTION"]))
    if len(pairs) > max_size:
        pairs = pairs[:max_size]
    return pairs

_cameo_cache: Dict[str, Tuple[str,str,float]] = {}
def _key_text(text: str, limit: int) -> str:
    s = (text or "")[:limit].encode("utf-8", "ignore")
    return hashlib.sha1(s).hexdigest()

def infer_cameo_code_fast_tree(
    text: str,
    ev_df: pd.DataFrame,
    zs_pipe,
    root_code: str = "1",
    refine_depth: int = 2,
    refine_max: int = 60,
    batch_len_limit: int = 1000,
    hypothesis_template: Optional[str] = None,
    early_stop: float = 0.60,
    coarse_codes: Optional[List[str]] = None
) -> Tuple[str, str, float]:
    sample = (text or "").strip()
    if not sample:
        return "", "", 0.0
    sample = sample[:batch_len_limit]

    k = _key_text(sample, limit=batch_len_limit)
    if k in _cameo_cache:
        return _cameo_cache[k]

    current = root_code
    if coarse_codes:
        root_guess, root_score = coarse_pick_root(sample, coarse_codes, ev_df, zs_pipe, hypothesis_template)
        if root_guess:
            current = root_guess
            if root_score >= early_stop:
                result = (current, f"[{current}] (coarse)", root_score)
                _cameo_cache[k] = result
                return result

    best_score = 0.0
    best_label = ""
    for _ in range(2):
        cand = descendant_codes(ev_df, current, max_depth_extra=refine_depth, max_candidates=refine_max)
        clabels = [f"[{c}] {d}" if d else f"[{c}]" for c, d in cand]
        lab, scr = _zs_best_label(sample, clabels, zs_pipe, hypothesis_template)
        if lab is None:
            break
        m = re.match(r"\[(.*?)\]", lab)
        code = m.group(1) if m else current
        best_label, best_score = lab, scr
        if scr >= early_stop or code == current:
            current = code
            break
        current = code

    result = (current, best_label, best_score)
    _cameo_cache[k] = result
    return result

def infer_cameo_code_subset(
    text: str,
    subset_pairs: List[Tuple[str,str]],
    zs_pipe,
    batch_len_limit: int = 1000,
    hypothesis_template: Optional[str] = None
) -> Tuple[str, str, float]:
    sample = (text or "").strip()
    if not sample:
        return "", "", 0.0
    sample = sample[:batch_len_limit]

    labels = [f"[{c}] {d}" if d else f"[{c}]" for c,d in subset_pairs]
    lab, scr = _zs_best_label(sample, labels, zs_pipe, hypothesis_template)
    if not lab:
        return "", "", 0.0
    m = re.match(r"\[(.*?)\]", lab)
    code = m.group(1) if m else ""
    return code, lab, scr

def infer_cameo_for_dataframe(
    df: pd.DataFrame,
    ev_df: pd.DataFrame,
    ev_desc: Dict[str, str],
    gs_map: Dict[str, float],
    zs_pipe,
    root_code: str,
    refine_depth: int,
    refine_max: int,
    batch_len_limit: int,
    hypothesis_template: Optional[str],
    early_stop: float,
    coarse_codes: Optional[List[str]],
    subset_pairs: Optional[List[Tuple[str,str]]] = None
) -> pd.DataFrame:
    cameo_codes, cameo_descs, golds, zs_labels, zs_scores, methods = [], [], [], [], [], []
    use_subset = bool(subset_pairs)

    for _, row in df.iterrows():
        t = str(row.get("text", "") or "").strip()
        if not t:
            cameo_codes.append("")
            cameo_descs.append("(unknown)")
            golds.append(0.0)
            zs_labels.append("")
            zs_scores.append(0.0)
            methods.append("none")
            continue

        if use_subset:
            code, zslab, zsscore = infer_cameo_code_subset(
                t, subset_pairs, zs_pipe,
                batch_len_limit=batch_len_limit,
                hypothesis_template=hypothesis_template
            )
            method = "subset-flat"
        else:
            code, zslab, zsscore = infer_cameo_code_fast_tree(
                t, ev_df, zs_pipe,
                root_code=root_code,
                refine_depth=refine_depth,
                refine_max=refine_max,
                batch_len_limit=batch_len_limit,
                hypothesis_template=hypothesis_template,
                early_stop=early_stop,
                coarse_codes=coarse_codes
            )
            method = "zero-shot-hier-fast"

        desc, gold = lookup_with_prefix_fallback(code, ev_desc, gs_map)
        cameo_codes.append(code)
        cameo_descs.append(desc)
        golds.append(gold)
        zs_labels.append(zslab)
        zs_scores.append(zsscore)
        methods.append(method)

    out = df.copy()
    out["cameo_code"] = cameo_codes
    out["cameo_description"] = cameo_descs
    out["goldstein"] = golds
    out["zs_label"] = zs_labels
    out["zs_score"] = zs_scores
    out["method"] = methods
    return out

# ---------------------------
# Sentometric: helpers
# ---------------------------
def ensure_nltk():
    try:
        nltk.data.find("tokenizers/punkt")
    except LookupError:
        nltk.download("punkt", quiet=True)
    try:
        nltk.data.find("sentiment/vader_lexicon.zip")
    except LookupError:
        nltk.download("vader_lexicon", quiet=True)

# Quote / boilerplate stripping
QUOTE_RX = re.compile(r'^[\"“”„«»].+?[\"“”„«»]$')
SPEAKER_RX = re.compile(r'^\s*(a (spus|declarat)|said|stated|declara|potrivit)\b', re.I)

def strip_quotes_boiler(text: str) -> str:
    lines = [ln.strip() for ln in text.splitlines()]
    keep = []
    for ln in lines:
        if not ln:
            continue
        if QUOTE_RX.match(ln):
            continue
        if SPEAKER_RX.match(ln):
            continue
        if len(ln) < 25 and ln.lower() in {"share", "comments", "accept cookies", "cookie", "citește și"}:
            continue
        keep.append(ln)
    return "\n".join(keep)

# Sentence weighting (lead + monument focus + vandalism boost)
FOCUS_KWS = [
    "statu", "monument", "memorial", "bust", "placă", "placa", "piatră",
    "szobor", "emlékmű", "emlekmu", "plăcu", "demolar", "dezafect", "topo",
    "vandal", "deface", "graffiti", "reamplas", "soclu", "inaugur", "protest"
]
FOCUS_NEG = [
    # Romanian/Hungarian vandalism/damage triggers (stems included)
    "vandal", "distrus", "distrug", "spart", "profan", "desecr",
    "răsturn", "rasturn", "smuls", "furat", "pângăr", "pangar",
    "megrong", "ledönt", "leönt", "felgyújt", "felgyujt", "rongál", "rongal",
    "distrusă", "distrusa"
]

def _weights_for_sentences(sents: List[str], lead_bonus=1.5, focus_bonus=2.0, vandal_bonus=3.0):
    w = []
    for i, s in enumerate(sents):
        base = 1.0
        if i < 5:
            base *= lead_bonus
        low = s.lower()
        if any(k in low for k in FOCUS_KWS):
            base *= focus_bonus
        if any(k in low for k in FOCUS_NEG):
            base *= vandal_bonus
        w.append(base)
    return np.asarray(w, dtype=float)

def _weighted_mean(vals: List[float], w: np.ndarray) -> float:
    if not vals:
        return 0.0
    w = w[:len(vals)]
    v = np.asarray(vals, dtype=float)[:len(w)]
    return float((v * w).sum() / (w.sum() + 1e-9))

# Label rule: symmetric or asymmetric neutral window
def label_from_signed(mean_signed: float, neutral_band: float, asym_neg: Optional[float], asym_pos: Optional[float]) -> str:
    if asym_neg is not None and asym_pos is not None:
        if -asym_neg < mean_signed < asym_pos:
            return "neutral"
        return "positive" if mean_signed >= asym_pos else "negative"
    if abs(mean_signed) < neutral_band:
        return "neutral"
    return "positive" if mean_signed > 0 else "negative"

# Alignment helper
def clip_text_for_sentiment(text: str, align_by: str, max_sents: int, sample_chars: int) -> str:
    t = (text or "").strip()
    if not t:
        return ""
    if align_by == "chars":
        return t[:sample_chars]
    sents = sent_tokenize(t) or [t]
    return " ".join(sents[:max_sents])

# --- MarianMT (optional) for VADER translation ---
_MT_CACHE: Dict[str, Tuple[AutoTokenizer, AutoModelForSeq2SeqLM, dict]] = {}
_MT_FAIL = set()

def _pick_device_for_mt(user_device: int):
    if torch.backends.mps.is_available():
        return "mps"
    if user_device is not None and user_device >= 0 and torch.cuda.is_available():
        return f"cuda:{user_device}"
    return "cpu"

def build_mt_model(direction: str, device: int, max_length: int = 512):
    model_map = {
        "ro-en": "Helsinki-NLP/opus-mt-ro-en",
        "hu-en": "Helsinki-NLP/opus-mt-hu-en",
    }
    name = model_map.get(direction)
    if not name:
        return None, None, None
    try:
        tok = AutoTokenizer.from_pretrained(name)
        mdl = AutoModelForSeq2SeqLM.from_pretrained(name)
        dev_str = _pick_device_for_mt(device)
        mdl = mdl.to(dev_str)
        gen_kwargs = dict(max_length=max_length, num_beams=1)
        print(f"[info] MT {direction} device: {dev_str}")
        return tok, mdl, gen_kwargs
    except Exception as e:
        print(f"[warn] MT model load failed ({direction}): {e}")
        return None, None, None

def _mt_chunk_translate(text: str, tok, mdl, gen_kwargs, chunk_chars: int = 600):
    if not text:
        return text
    pieces = []
    start = 0
    L = len(text)
    while start < L:
        chunk = text[start:start+chunk_chars]
        start += chunk_chars
        try:
            inputs = tok(chunk, return_tensors="pt", truncation=True, max_length=gen_kwargs.get("max_length", 512))
            inputs = {k: v.to(mdl.device) for k, v in inputs.items()}
            outputs = mdl.generate(**inputs, **gen_kwargs)
            out = tok.batch_decode(outputs, skip_special_tokens=True)[0]
            pieces.append(out)
        except Exception:
            pieces.append(chunk)
    return " ".join(pieces)

def maybe_translate_for_vader(text: str, mode: str, device: int) -> str:
    mode = (mode or "off").lower()
    if mode == "off":
        return text

    target = None
    if mode in ("ro-en", "hu-en"):
        target = mode
    elif mode == "auto":
        lang = None
        if HAVE_LANGDETECT:
            try:
                lang = detect(text[:1000])
            except Exception:
                lang = None
        if lang == "ro" or re.search(r"[ăâîșţțș]", text.lower()):
            target = "ro-en"
        elif lang == "hu":
            target = "hu-en"
        else:
            return text
    else:
        return text

    if target in _MT_FAIL:
        return text
    if target not in _MT_CACHE:
        tok, mdl, gen_kwargs = build_mt_model(target, device=device, max_length=512)
        if tok is None:
            _MT_FAIL.add(target)
            return text
        _MT_CACHE[target] = (tok, mdl, gen_kwargs)

    tok, mdl, gen_kwargs = _MT_CACHE[target]
    try:
        return _mt_chunk_translate(text, tok, mdl, gen_kwargs, chunk_chars=600)
    except Exception:
        return text

# --- Zero-shot SENTIMENT (replaces old "HF" sentiment) ---
ZS_SENT_LABELS = ["negative", "neutral", "positive"]
ZS_HYPOTHESIS = "Această propoziție exprimă un sentiment {}."

def build_zs_sentiment(model_name: str, device: int):
    # reuse zero-shot pipeline with same NLI model
    return pipeline("zero-shot-classification", model=model_name, device=device)

def zs_sent_scores_for_sentences(zs_pipe, sents: List[str]) -> List[Dict[str,float]]:
    """
    Returns per-sentence dicts: {"negative": p, "neutral": p, "positive": p}
    Uses multi_label=True to get independent scores.
    """
    results: List[Dict[str,float]] = []
    if not sents:
        return results
    batch = 12
    for i in range(0, len(sents), batch):
        chunk = sents[i:i+batch]
        out = zs_pipe(
            chunk,
            candidate_labels=ZS_SENT_LABELS,
            hypothesis_template=ZS_HYPOTHESIS,
            multi_label=True,
        )
        if isinstance(out, dict):
            out = [out]
        for r in out:
            d = {lab.lower(): sc for lab, sc in zip(r["labels"], r["scores"])}
            results.append({k: float(d.get(k, 0.0)) for k in ZS_SENT_LABELS})
    return results

def signed_from_pos_neg(d: Dict[str,float]) -> float:
    return float(d.get("positive", 0.0) - d.get("negative", 0.0))

# Domain prior families (hostile/destructive)
NEG_FAMS = ("145", "171", "1712", "18", "182")

# --- VADER augmentation for RO/HU news about monuments/vandalism ---
RO_HU_VADER_LEXICON = {
    # Romanian vandalism / destruction (stems and inflections)
    "vandalizat": -3.5, "vandalizare": -3.5, "vandalism": -3.5,
    "distrus": -3.4, "distrusa": -3.4, "distrugere": -3.4, "spart": -2.8,
    "profanat": -3.5, "pangarit": -3.2, "rasturnat": -3.0,
    "smuls": -2.8, "furat": -2.8, "desecrat": -3.4, "incendiat": -3.6,
    "agresiune": -2.8, "violente": -3.0, "reprimare": -2.6,

    # Hungarian vandalism / destruction
    "megrongált": -3.5, "megrongalt": -3.5, "rongálás": -3.5, "rongalas": -3.5,
    "ledöntött": -3.4, "ledontott": -3.4, "leöntött": -2.8, "leontott": -2.8,
    "felgyújtott": -3.6, "felgyujtott": -3.6, "támadás": -2.8, "eroszak": -3.0,

    # Protest / tension (slightly negative)
    "protest": -1.2, "proteste": -1.2, "scandal": -1.6,
    "tensiune": -1.2, "conflict": -1.6,

    # Heritage/monument praise (weak positive so it won't dominate)
    "monument": +0.6, "patrimoniu": +0.6, "istoric": +0.4, "cultural": +0.4,
    "inaugurare": +0.8, "omagiere": +0.8, "sarbatoare": +0.6,
    "szobor": +0.6, "emlékmű": +0.6, "emlekmu": +0.6,
}

# Vandalism triggers to detect context (for penalty)
VADER_VANDAL_TRIGGERS = {
    "vandal", "distrug", "distrus", "profan", "pangar", "incend",
    "rasturn", "smuls", "furat", "violente", "agresiune",
    "megrong", "ledont", "leont", "felgyujt", "rongal"
}

def build_vader_analyzer(neg_scale: float = 1.15, pos_scale: float = 0.60):
    """
    Returns (analyzer, rescale_func), with:
      - analyzer: VADER SentimentIntensityAnalyzer augmented for RO/HU domain
      - rescale_func(x): asymmetric squash (amplify negatives, shrink positives)
    """
    analyzer = SentimentIntensityAnalyzer()
    analyzer.lexicon.update(RO_HU_VADER_LEXICON)

    def rescale_signed(x: float) -> float:
        return (x * pos_scale) if x >= 0 else (x * neg_scale)

    return analyzer, rescale_signed

def compute_sentometric(
    df: pd.DataFrame,
    zs_sent_pipe,            # zero-shot sentiment pipeline (mDeBERTa-based)
    max_sents: int,
    align_by: str,
    sample_chars: int,
    neutral_band: float,
    vader_translate: str,
    mt_device: int,
    asym_neg: Optional[float],
    asym_pos: Optional[float],
    use_domain_prior: bool,
    domain_prior_weight: float,
    use_vader: bool = True
) -> pd.DataFrame:
    ensure_nltk()

    # Build VADER analyzer once (augmented) if enabled
    vader_an = vader_rescale = None
    if use_vader:
        vader_an, vader_rescale = build_vader_analyzer(
            neg_scale=1.15,  # amplify negatives by 15%
            pos_scale=0.60   # shrink positives to 60%
        )

    v_tone, v_mag, v_lbl = [], [], []
    hf_tone, hf_mag, hf_lbl = [], [], []  # "HF" columns hold Zero-Shot sentiment

    for _, row in df.iterrows():
        raw = str(row.get("text", "") or "").strip()
        if not raw:
            v_tone.append(None); v_mag.append(None); v_lbl.append(None)
            hf_tone.append(None); hf_mag.append(None); hf_lbl.append(None)
            continue

        # Clean quotes/boilerplate
        txt = strip_quotes_boiler(raw)
        sample = clip_text_for_sentiment(txt, align_by=align_by, max_sents=max_sents, sample_chars=sample_chars)
        sents = sent_tokenize(sample) or [sample]
        w = _weights_for_sentences(sents)

        # 1) VADER (optional, with translation + augmentation + rescale + context penalty)
        if vader_an is not None:
            v_sample = maybe_translate_for_vader(sample, vader_translate, device=mt_device)
            v_sents = sent_tokenize(v_sample) or [v_sample]

            v_comp_raw  = [vader_an.polarity_scores(s)["compound"] for s in v_sents]
            v_comp      = [vader_rescale(c) for c in v_comp_raw]  # asymmetric rescale

            wv      = _weights_for_sentences(v_sents)
            v_mean_signed = _weighted_mean(v_comp, wv)
            v_mean_abs    = _weighted_mean([abs(c) for c in v_comp], wv)

            # Vandalism context penalty
            if any(t in sample.lower() for t in VADER_VANDAL_TRIGGERS) and v_mean_signed > -0.05:
                v_mean_signed -= 0.20

            v_label = label_from_signed(v_mean_signed, neutral_band, asym_neg, asym_pos)
            v_tone_val = round(100.0 * v_mean_signed, 3)
            v_mag_val  = round(100.0 * v_mean_abs, 3)

            # domain prior (only on neutral/positive)
            if use_domain_prior and v_label != "negative":
                code = str(row.get("cameo_code") or "")
                if any(code.startswith(f) for f in NEG_FAMS):
                    v_tone_val -= 100.0 * float(domain_prior_weight)
                    v_label = "negative"

            v_tone.append(v_tone_val); v_mag.append(v_mag_val); v_lbl.append(v_label)
        else:
            v_tone.append(None); v_mag.append(None); v_lbl.append(None)

        # 2) ZERO-SHOT SENTIMENT (better for news, replaces old HF review model)
        if zs_sent_pipe is None:
            hf_tone.append(None); hf_mag.append(None); hf_lbl.append(None)
            continue

        try:
            scores_list = zs_sent_scores_for_sentences(zs_sent_pipe, sents)
        except Exception:
            scores_list = []

        signed_vals = [signed_from_pos_neg(d) for d in scores_list] if scores_list else []
        if not signed_vals:
            hf_tone.append(None); hf_mag.append(None); hf_lbl.append(None)
            continue

        # vandalism override: if article contains vandalism triggers and mean is "neutralish", push negative
        article_low = sample.lower()
        vandal_hit = any(k in article_low for k in FOCUS_NEG)

        # weighted means
        mean_signed = _weighted_mean(signed_vals, w)
        mean_abs    = _weighted_mean([abs(x) for x in signed_vals], w)

        # gentle override before labeling
        if vandal_hit and mean_signed > -0.05:
            mean_signed -= 0.25  # nudge into negative zone

        label = label_from_signed(mean_signed, neutral_band, asym_neg, asym_pos)
        tone_val = round(100.0 * mean_signed, 3)
        mag_val  = round(100.0 * mean_abs, 3)

        # domain prior: if still neutral/positive but destructive CAMEO, nudge negative
        if use_domain_prior and label != "negative":
            code = str(row.get("cameo_code") or "")
            if any(code.startswith(f) for f in NEG_FAMS):
                tone_val -= 100.0 * float(domain_prior_weight)
                label = "negative"

        hf_tone.append(tone_val); hf_mag.append(mag_val); hf_lbl.append(label)

    out = df.copy()
    out["vader_tone"] = v_tone
    out["vader_magnitude"] = v_mag
    out["vader_label"] = v_lbl
    # NOTE: Below "hf_*" columns are produced by Zero-Shot sentiment:
    out["hf_tone"] = hf_tone
    out["hf_magnitude"] = hf_mag
    out["hf_label"] = hf_lbl
    return out

# ---------------------------
# main / CLI
# ---------------------------
def parse_args():
    ap = argparse.ArgumentParser(
        description="Scrape texts, infer CAMEO (subset or fast tree), join Goldstein, and compute news-friendly sentometrics."
    )
    # Scrape
    ap.add_argument("--excel", required=True)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--url-col", type=int, required=True)
    ap.add_argument("--one-based", action="store_true")
    ap.add_argument("--timeout", type=int, default=40)

    # CAMEO + Goldstein
    ap.add_argument("--eventcodes", required=True)
    ap.add_argument("--goldstein", required=True)

    # Zero-shot model & speed (also used for ZS sentiment)
    ap.add_argument("--zs-model", default="MoritzLaurer/mDeBERTa-v3-base-mnli-xnli")
    ap.add_argument("--device", type=int, default=-1)  # -1 CPU (safe)
    ap.add_argument("--root", default="1")
    ap.add_argument("--refine-depth", type=int, default=2)
    ap.add_argument("--refine-max", type=int, default=60)
    ap.add_argument("--zs-early-stop", type=float, default=0.60)
    ap.add_argument("--zs-sample-chars", type=int, default=1000)
    ap.add_argument("--zs-coarse", default="1,14,18")
    ap.add_argument("--hypothesis", default="Această propoziție este despre {}.")

    # Subset (statues/monuments)
    ap.add_argument("--zs-subset-codes",
                    default="14,141,145,18,171,1712,017,051,055,057,111,112,114,140,142,143,144,175,182,1823")
    ap.add_argument("--zs-subset-keywords",
                    default="statue,monument,memorial,heritage,unveil,inaugur,commemor,remove,topple,vandal,deface,demolish,rename,protest")
    ap.add_argument("--zs-subset-depth", type=int, default=2)
    ap.add_argument("--zs-subset-max", type=int, default=120)
    ap.add_argument("--zs-subset-disable", action="store_true")

    # Sentiment (ZS + VADER + neutral reduction + domain prior)
    ap.add_argument("--sent-align-by", choices=["sents","chars"], default="sents")
    ap.add_argument("--sent-max-sents", type=int, default=120)
    ap.add_argument("--sent-sample-chars", type=int, default=2000)
    ap.add_argument("--sent-neutral-band", type=float, default=0.05)
    ap.add_argument("--sent-asym-neg", type=float, default=0.03)
    ap.add_argument("--sent-asym-pos", type=float, default=0.10)

    # VADER translation for RO/HU (optional; ZS runs on original RO/HU)
    ap.add_argument("--vader-translate", choices=["off","auto","ro-en","hu-en"], default="off")
    ap.add_argument("--mt-device", type=int, default=-1, help="MT device hint (CPU by default; MPS auto-used if available).")

    # Domain prior
    ap.add_argument("--sent-domain-prior", choices=["on","off"], default="on")
    ap.add_argument("--sent-domain-prior-weight", type=float, default=0.12,
                    help="How much to nudge neutral/positive → negative when CAMEO is a negative family (0..1).")

    # Output
    ap.add_argument("--out", default="scraped_with_cameo_sentiment.csv")
    return ap.parse_args()

def main():
    args = parse_args()

    # URLs from Excel
    _ = build_trafilatura_config()
    urls, _, _ = read_urls_with_hyperlinks(args.excel, args.sheet, args.url_col, args.one_based)
    if not urls:
        print("[warn] No URLs found in that column.", file=sys.stderr)

    # Scrape
    scraped_df = scrape_urls(urls, timeout=args.timeout)

    # Load CAMEO + Goldstein
    ev_df, ev_desc, gs_map = load_cameo_tables(args.eventcodes, args.goldstein)

    # Zero-shot model (used both for CAMEO and SENTIMENT)
    zs_pipe = build_zs(args.zs_model, args.device)
    zs_sent_pipe = build_zs_sentiment(args.zs_model, args.device)

    # Subset (if enabled)
    subset_pairs = None
    if not args.zs_subset_disable:
        seeds = _normalize_kw_list(args.zs_subset_codes)
        kws = _normalize_kw_list(args.zs_subset_keywords)
        subset_pairs = subset_from_keywords_and_seeds(
            ev_df,
            seed_codes=seeds,
            keywords=kws,
            depth=args.zs_subset_depth,
            max_size=args.zs_subset_max
        )
        if subset_pairs:
            print(f"[info] subset enabled: {len(subset_pairs)} codes in candidate pool")
        else:
            print("[info] subset produced 0 codes; falling back to fast tree")

    # Infer CAMEO (+Goldstein)
    coarse_list = [c.strip() for c in args.zs_coarse.split(",")] if args.zs_coarse.strip() else None
    df_cameo = infer_cameo_for_dataframe(
        scraped_df, ev_df, ev_desc, gs_map, zs_pipe,
        root_code=args.root,
        refine_depth=args.refine_depth,
        refine_max=args.refine_max,
        batch_len_limit=args.zs_sample_chars,
        hypothesis_template=args.hypothesis,
        early_stop=args.zs_early_stop,
        coarse_codes=coarse_list,
        subset_pairs=subset_pairs
    )

    # Sentiment (Zero-Shot + augmented VADER)
    final_df = compute_sentometric(
        df_cameo,
        zs_sent_pipe=zs_sent_pipe,
        max_sents=args.sent_max_sents,
        align_by=args.sent_align_by,
        sample_chars=args.sent_sample_chars,
        neutral_band=args.sent_neutral_band,
        vader_translate=args.vader_translate,
        mt_device=args.mt_device,
        asym_neg=args.sent_asym_neg,
        asym_pos=args.sent_asym_pos,
        use_domain_prior=(args.sent_domain_prior == "on"),
        domain_prior_weight=args.sent_domain_prior_weight,
        use_vader=True
    )

    final_df.to_csv(args.out, index=False, encoding="utf-8")
    print("\nSaved:", args.out)

if __name__ == "__main__":
    main()

# python step1.py \
#   --excel "/path/Articole_de_presa_proteste_statui_CJ-N.xlsx" \
#   --sheet Sheet1 --url-col 6 --one-based --timeout 200 \
#   --eventcodes "/path/CAMEO.eventcodes.txt" \
#   --goldstein "/path/CAMEO.goldsteinscale.txt" \
#   --zs-model MoritzLaurer/mDeBERTa-v3-base-mnli-xnli \
#   --zs-subset-codes 017,051,055,057,111,112,114,14,140,141,142,143,144,145,171,1712,175,18,182,1823 \
#   --zs-subset-keywords statue,monument,memorial,heritage,unveil,inaugur,commemor,remove,topple,vandal,deface,rename,protest \
#   --zs-sample-chars 3000 --zs-early-stop 0.8 \
#   --sent-align-by sents --sent-max-sents 120 \
#   --sent-neutral-band 0.05 --sent-asym-neg 0.03 --sent-asym-pos 0.10 \
#   --sent-domain-prior on --sent-domain-prior-weight 0.12 \
#   --vader-translate off \
#   --out scraped_with_cameo_sentiment_newsZS.csv
